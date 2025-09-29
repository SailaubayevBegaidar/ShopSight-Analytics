from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
import os
import pandas as pd

def export_to_excel(dataframes_dict, filename):
    EXPORT_DIR = "exports"
    os.makedirs(EXPORT_DIR, exist_ok=True)
    filepath = os.path.join(EXPORT_DIR, filename)

    # Сохраняем все датафреймы в Excel
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        # save() не нужен, файл сохраняется автоматически

    # Форматирование через openpyxl
    wb = load_workbook(filepath)
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        total_rows += ws.max_row - 1  # минус заголовок

        # Заморозка первой строки и колонки
        ws.freeze_panes = "B2"

        # Автофильтры
        ws.auto_filter.ref = ws.dimensions

        # Последняя колонка
        last_col = ws.max_column
        col_letter = ws.cell(row=1, column=last_col).column_letter

        # Проверяем, числовая ли колонка
        sample_values = [ws.cell(row=r, column=last_col).value for r in range(2, min(ws.max_row+1, 4))]
        if all(isinstance(val, (int, float)) for val in sample_values if val is not None):
            rule = ColorScaleRule(
                start_type="min", start_color="FFAA0000",
                mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                end_type="max", end_color="FF00AA00"
            )
            ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

    wb.save(filepath)
    print(f"Created file {filename}, {len(wb.sheetnames)} sheets, {total_rows} rows")


if __name__ == "__main__":
    # Пример датафреймов
    df1 = pd.DataFrame({
        "Month": ["2025-01", "2025-02", "2025-03"],
        "Orders": [120, 150, 130],
        "Revenue": [5000, 6200, 5800]
    })

    df2 = pd.DataFrame({
        "Category": ["Electronics", "Books", "Clothing"],
        "Sales": [300, 250, 180],
        "Revenue": [15000, 12000, 9000]
    })

    dataframes = {
        "Orders_per_Month": df1,
        "Top_Categories": df2
    }

    export_to_excel(dataframes, "sales_report.xlsx")
